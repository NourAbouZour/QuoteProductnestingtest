# doc_generator.py
"""
DEVNOTE: Windows-friendly PDF generation with pluggable renderers
================================================================

This module uses a pluggable renderer architecture for PDF generation:
1. PRIMARY: Playwright + Chromium (Windows-friendly, no external dependencies)
2. FALLBACK: WeasyPrint (if Cairo/Pango available)  
3. FALLBACK: wkhtmltopdf (if binary in PATH)

Setup for Windows:
1. pip install playwright
2. playwright install chromium

The HTML template uses CSS paged media with proper page breaks, margins, 
and repeating table headers for professional PDF output.

Author: ChatGPT for Nour
Python: 3.10+
"""

from __future__ import annotations
import os
import sys
import json
import tempfile
import subprocess
from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime
import base64
import re
import io

# Image processing imports
try:
    from PIL import Image, ImageCms
    import numpy as np
    _HAVE_PIL = True
except ImportError:
    _HAVE_PIL = False
    print("Warning: PIL/Pillow not available. Image recoloring will be disabled.")

# SVG processing imports
try:
    import svgwrite
    _HAVE_SVGWRITE = True
except ImportError:
    _HAVE_SVGWRITE = False
    print("Warning: svgwrite not available. SVG conversion will be disabled.")

# DXF processing imports
try:
    import ezdxf
    _HAVE_EZDXF = True
except ImportError:
    _HAVE_EZDXF = False
    print("Warning: ezdxf not available. DXF processing will be disabled.")

# PDF renderer imports
try:
    from playwright.sync_api import sync_playwright
    _HAVE_PLAYWRIGHT = True
except ImportError:
    _HAVE_PLAYWRIGHT = False

try:
    from weasyprint import HTML
    _HAVE_WEASY = True
except (ImportError, OSError):
    _HAVE_WEASY = False
    print("Warning: WeasyPrint not available. Will use Playwright or wkhtmltopdf.")

try:
    import pdfkit
    _HAVE_WKHTMLTOPDF = True
except ImportError:
    _HAVE_WKHTMLTOPDF = False

from pydantic import BaseModel, Field, validator

# Third-party deps:
#   pip install jinja2 pydantic openpyxl pillow playwright
from jinja2 import Environment, BaseLoader, select_autoescape
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


# ---- Image Recoloring Configuration ----
NEAR_BLACK_MAX = 40          # <= this => treat as background (black/dark gray)
GRID_MAX       = 90          # <= this => also treat as background (kills gray grid)
LASER_MIN      = 200         # >= this on all channels => laser (white)
VG_R_MIN       = 160         # red min (vgroove)
VG_G_MAX       = 90          # green max for "redness"
VG_B_MAX       = 90          # blue max for "redness"

# Color definitions for different element types
LASER_COLOR    = (0, 0, 0)       # black - general part lines
VGROOVE_COLOR  = (0, 160, 0)     # green - v-groove lines
CIRCLE_COLOR   = (255, 0, 0)     # red - circles
BENDING_COLOR  = (0, 0, 255)     # blue - bending lines
BACKGROUND     = (255, 255, 255) # white

def png_to_svg_data_url(png_data_url: str, max_width: int = 80, max_height: int = 65) -> str:
    """Convert PNG image to SVG with smooth vector lines, optimized for PDF table cells."""
    if not _HAVE_PIL or not _HAVE_SVGWRITE:
        return png_data_url  # Return original if dependencies not available
    
    try:
        # 1) Decode PNG
        m = re.match(r"^data:image/(png|jpeg);base64,(.+)$", png_data_url, flags=re.I|re.S)
        if not m:
            return png_data_url
        
        raw = base64.b64decode(m.group(2))
        im = Image.open(io.BytesIO(raw)).convert("RGBA")
        
        # 2) Get image dimensions and scale
        width, height = im.size
        scale_x = max_width / width
        scale_y = max_height / height
        scale = min(scale_x, scale_y, 1.0)  # Don't scale up, only down
        
        svg_width = int(width * scale)
        svg_height = int(height * scale)
        
        # 3) Convert to numpy array for processing
        arr = np.array(im, dtype=np.uint8)
        rgb = arr[..., :3].astype(np.uint16)
        a = arr[..., 3:4]
        
        R, G, B = rgb[..., 0], rgb[..., 1], rgb[..., 2]
        
        # 4) Create masks for different line types
        # Background mask
        m_bg = (R <= NEAR_BLACK_MAX) & (G <= NEAR_BLACK_MAX) & (B <= NEAR_BLACK_MAX)
        m_grid = (R <= GRID_MAX) & (G <= GRID_MAX) & (B <= GRID_MAX)
        m_alpha_bg = (a[..., 0] <= 10)
        background_mask = m_bg | m_grid | m_alpha_bg
        
        # V-groove mask (reddish areas)
        vg_mask = (R >= VG_R_MIN) & (G <= VG_G_MAX) & (B <= VG_B_MAX)
        
        # Laser mask (bright/white areas)
        # Also catch medium-bright areas that are likely general part lines
        laser_mask = ((R >= LASER_MIN) & (G >= LASER_MIN) & (B >= LASER_MIN)) | \
                     ((R >= 100) & (G >= 100) & (B >= 100) & (R <= 200) & (G <= 200) & (B <= 200))
        
        # Circle mask (detect circular patterns)
        circle_mask = (R >= 140) & (R <= 200) & (G >= 100) & (G <= 160) & (B >= 100) & (B <= 160)
        
        # Bending mask (detect bending lines)
        bending_mask = (R >= 120) & (R <= 180) & (G >= 120) & (G <= 180) & (B >= 180) & (B <= 250)
        
        # 5) Create SVG
        dwg = svgwrite.Drawing(size=(f"{svg_width}px", f"{svg_height}px"))
        
        # Add white background
        dwg.add(dwg.rect(insert=(0, 0), size=("100%", "100%"), fill="white"))
        
        # 6) Convert pixels to vector paths using line detection
        # For general part lines (default black)
        laser_pixels = np.where(((laser_mask | (~vg_mask & ~circle_mask & ~bending_mask)) & ~background_mask))
        if len(laser_pixels[0]) > 0:
            laser_lines = detect_and_connect_lines(laser_pixels, scale, width, height)
            for line in laser_lines:
                dwg.add(dwg.line(
                    start=(line[0], line[1]),
                    end=(line[2], line[3]),
                    stroke="black",
                    stroke_width=max(1, int(20.0 * scale))
                ))
        
        # For v-groove lines (green, smooth)
        vg_pixels = np.where(vg_mask & ~background_mask)
        if len(vg_pixels[0]) > 0:
            vg_lines = detect_and_connect_lines(vg_pixels, scale, width, height)
            for line in vg_lines:
                dwg.add(dwg.line(
                    start=(line[0], line[1]),
                    end=(line[2], line[3]),
                    stroke="green",
                    stroke_width=max(1, int(20.0 * scale))
                ))
                
        # For circle lines (red, smooth)
        circle_pixels = np.where(circle_mask & ~background_mask)
        if len(circle_pixels[0]) > 0:
            circle_lines = detect_and_connect_lines(circle_pixels, scale, width, height)
            for line in circle_lines:
                dwg.add(dwg.line(
                    start=(line[0], line[1]),
                    end=(line[2], line[3]),
                    stroke="red",
                    stroke_width=max(1, int(20.0 * scale))
                ))
                
        # For bending lines (blue, smooth)
        bending_pixels = np.where(bending_mask & ~background_mask)
        if len(bending_pixels[0]) > 0:
            bending_lines = detect_and_connect_lines(bending_pixels, scale, width, height)
            for line in bending_lines:
                dwg.add(dwg.line(
                    start=(line[0], line[1]),
                    end=(line[2], line[3]),
                    stroke="blue",
                    stroke_width=max(1, int(20.0 * scale))
                ))
        
        # 7) Convert SVG to data URL
        svg_string = dwg.tostring()
        svg_b64 = base64.b64encode(svg_string.encode('utf-8')).decode('ascii')
        return f"data:image/svg+xml;base64,{svg_b64}"
        
    except Exception as e:
        print(f"Warning: PNG to SVG conversion failed: {e}")
        return png_data_url  # Return original on error

def detect_and_connect_lines(pixel_coords, scale, width, height):
    """
    Detect and connect line segments from pixel coordinates.
    Returns a list of line segments as (x1, y1, x2, y2) tuples.
    """
    if len(pixel_coords[0]) == 0:
        return []
    
    # Try OpenCV-based detection first (if available)
    try:
        return detect_lines_opencv(pixel_coords, scale, width, height)
    except ImportError:
        # Fallback to pure Python implementation
        return detect_lines_pure_python(pixel_coords, scale, width, height)

def detect_lines_opencv(pixel_coords, scale, width, height):
    """Use OpenCV's Hough Line Transform for better straight line detection."""
    try:
        import cv2
        
        # Create a binary image from the pixel coordinates
        binary_image = np.zeros((height, width), dtype=np.uint8)
        y_coords, x_coords = pixel_coords
        binary_image[y_coords, x_coords] = 255
        
        # Apply morphological operations to clean up the image
        kernel = np.ones((2, 2), np.uint8)
        binary_image = cv2.morphologyEx(binary_image, cv2.MORPH_CLOSE, kernel)
        
        # Detect lines using Hough Transform
        lines = cv2.HoughLinesP(
            binary_image,
            rho=1,
            theta=np.pi/180,
            threshold=3,  # Lower threshold for thin lines
            minLineLength=3,
            maxLineGap=2
        )
        
        if lines is None:
            return []
        
        # Convert lines to our format
        line_segments = []
        for line in lines:
            x1, y1, x2, y2 = line[0]
            
            # Scale coordinates
            scaled_x1 = int(x1 * scale)
            scaled_y1 = int(y1 * scale)
            scaled_x2 = int(x2 * scale)
            scaled_y2 = int(y2 * scale)
            
            # Only add segments that are long enough
            if distance((scaled_x1, scaled_y1), (scaled_x2, scaled_y2)) >= 1:
                line_segments.append((scaled_x1, scaled_y1, scaled_x2, scaled_y2))
        
        return line_segments
        
    except ImportError:
        raise ImportError("OpenCV not available")
    except Exception as e:
        print(f"OpenCV line detection failed: {e}")
        raise

def detect_lines_pure_python(pixel_coords, scale, width, height):
    """
    Pure Python implementation for line detection.
    Falls back to this when OpenCV is not available.
    """
    # Convert to list of (x, y) coordinates
    points = list(zip(pixel_coords[1], pixel_coords[0]))  # Note: numpy.where returns (y, x)
    
    if len(points) <= 1:
        return []
    
    # Group connected pixels into line segments
    line_segments = []
    visited = set()
    
    for start_point in points:
        if start_point in visited:
            continue
        
        # Find connected component starting from this point
        component = find_connected_component(start_point, points, visited)
        if len(component) < 2:
            continue
        
        # Convert component to line segments
        segments = component_to_line_segments(component, scale)
        line_segments.extend(segments)
    
    return line_segments

def find_connected_component(start_point, all_points, visited):
    """Find all points connected to start_point using 8-connectivity."""
    component = []
    stack = [start_point]
    
    while stack:
        current = stack.pop()
        if current in visited:
            continue
        
        visited.add(current)
        component.append(current)
        
        # Check 8 neighbors
        x, y = current
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                if dx == 0 and dy == 0:
                    continue
                
                neighbor = (x + dx, y + dy)
                if neighbor in all_points and neighbor not in visited:
                    stack.append(neighbor)
    
    return component

def component_to_line_segments(component, scale):
    """Convert a connected component to simplified line segments."""
    if len(component) < 2:
        return []
    
    # Sort points to get a rough ordering along the line
    sorted_points = sort_points_along_line(component)
    
    # Simplify the path using Douglas-Peucker algorithm
    simplified = simplify_path(sorted_points, tolerance=1.0)
    
    # Convert to line segments
    segments = []
    for i in range(len(simplified) - 1):
        x1, y1 = simplified[i]
        x2, y2 = simplified[i + 1]
        
        # Scale coordinates
        scaled_x1 = int(x1 * scale)
        scaled_y1 = int(y1 * scale)
        scaled_x2 = int(x2 * scale)
        scaled_y2 = int(y2 * scale)
        
        # Only add segments that are long enough
        if distance((scaled_x1, scaled_y1), (scaled_x2, scaled_y2)) >= 1:
            segments.append((scaled_x1, scaled_y1, scaled_x2, scaled_y2))
    
    return segments

def sort_points_along_line(points):
    """Sort points to follow the line direction."""
    if len(points) <= 2:
        return points
    
    # Start with the leftmost point
    start = min(points, key=lambda p: p[0])
    sorted_points = [start]
    remaining = set(points) - {start}
    
    while remaining:
        current = sorted_points[-1]
        # Find the closest remaining point
        closest = min(remaining, key=lambda p: distance(current, p))
        sorted_points.append(closest)
        remaining.remove(closest)
    
    return sorted_points

def simplify_path(points, tolerance=1.0):
    """Simplify a path using Douglas-Peucker algorithm."""
    if len(points) <= 2:
        return points
    
    # Find the point with maximum distance from the line segment
    max_distance = 0
    max_index = 0
    
    start, end = points[0], points[-1]
    
    for i in range(1, len(points) - 1):
        dist = point_to_line_distance(points[i], start, end)
        if dist > max_distance:
            max_distance = dist
            max_index = i
    
    # If max distance is greater than tolerance, recursively simplify
    if max_distance > tolerance:
        left = simplify_path(points[:max_index + 1], tolerance)
        right = simplify_path(points[max_index:], tolerance)
        return left[:-1] + right  # Avoid duplicate point
    
    # Otherwise, return just the endpoints
    return [start, end]

def point_to_line_distance(point, line_start, line_end):
    """Calculate the distance from a point to a line segment."""
    px, py = point
    x1, y1 = line_start
    x2, y2 = line_end
    
    # Calculate the distance using the formula for point-to-line distance
    if x2 == x1:  # Vertical line
        return abs(px - x1)
    elif y2 == y1:  # Horizontal line
        return abs(py - y1)
    else:
        # General case
        A = y2 - y1
        B = x1 - x2
        C = x2 * y1 - x1 * y2
        return abs(A * px + B * py + C) / ((A * A + B * B) ** 0.5)

def distance(p1, p2):
    """Calculate Euclidean distance between two points."""
    return ((p1[0] - p2[0]) ** 2 + (p1[1] - p2[1]) ** 2) ** 0.5

def recolor_png_data_url(data_url: str) -> str:
    """Recolor PNG image: background to white, laser lines to black, v-groove lines to green, with enhanced borders and thicker lines."""
    if not _HAVE_PIL:
        return data_url  # Return original if PIL not available
    
    try:
        # 1) Decode
        m = re.match(r"^data:image/(png|jpeg);base64,(.+)$", data_url, flags=re.I|re.S)
        if not m:
            return data_url  # Return original if not valid data URL
        
        raw = base64.b64decode(m.group(2))
        im = Image.open(io.BytesIO(raw)).convert("RGBA")
        
        # Get original dimensions
        width, height = im.size
        
        # 2) Use original image without border expansion
        new_im = im.copy()
        
        # Convert to numpy array
        arr = np.array(new_im, dtype=np.uint8)
        rgb = arr[..., :3].astype(np.uint16)
        a = arr[..., 3:4]

        R, G, B = rgb[..., 0], rgb[..., 1], rgb[..., 2]

        # 3) Enhanced masks with better detection
        # background: near-black and grid-gray → white
        m_bg = (R <= NEAR_BLACK_MAX) & (G <= NEAR_BLACK_MAX) & (B <= NEAR_BLACK_MAX)
        m_grid = (R <= GRID_MAX) & (G <= GRID_MAX) & (B <= GRID_MAX)

        # v-groove: near red (more sensitive detection)
        m_vg = (R >= VG_R_MIN) & (G <= VG_G_MAX) & (B <= VG_B_MAX)

        # laser: bright/white strokes (more sensitive detection)
        # Also catch medium-bright areas that are likely general part lines
        m_laser = ((R >= LASER_MIN) & (G >= LASER_MIN) & (B >= LASER_MIN)) | \
                  ((R >= 100) & (G >= 100) & (B >= 100) & (R <= 200) & (G <= 200) & (B <= 200))

        # circles: detect based on shape characteristics (circular patterns)
        # Make circle detection more specific to avoid catching general part lines
        m_circle = (R >= 140) & (R <= 200) & (G >= 100) & (G <= 160) & (B >= 100) & (B <= 160)

        # bending lines: detect based on color characteristics
        # Bending lines are typically in a different color range
        m_bending = (R >= 120) & (R <= 180) & (G >= 120) & (G <= 180) & (B >= 180) & (B <= 250)

        # Anything transparent should be treated as background too
        m_alpha_bg = (a[..., 0] <= 10)

        # 4) Apply recoloring
        # Start with white canvas
        out = np.zeros_like(rgb)
        out[...] = BACKGROUND

        # Background mask (also kills grids)
        mask_bg_all = m_bg | m_grid | m_alpha_bg

        # Default: all non-background pixels become BLACK (general part lines)
        non_bg = ~mask_bg_all
        out[non_bg] = LASER_COLOR

        # Overrides in priority order (most specific first)
        # Bending → blue
        out[m_bending & non_bg] = BENDING_COLOR
        # Circles → red
        out[m_circle & non_bg & ~m_bending] = CIRCLE_COLOR
        # V-groove → green
        out[m_vg & non_bg & ~m_bending & ~m_circle] = VGROOVE_COLOR

        # 5) Thicken lines using morphological operations - MUCH BOLDER
        from scipy import ndimage
        import numpy as _np
        
        # Create masks for all line types
        laser_mask = (out == LASER_COLOR).all(axis=-1)
        vg_mask = (out == VGROOVE_COLOR).all(axis=-1)
        circle_mask = (out == CIRCLE_COLOR).all(axis=-1)
        bending_mask = (out == BENDING_COLOR).all(axis=-1)
        
        # Use a 21x21 square structuring element to enforce ~20px thickness
        _k20 = _np.ones((21, 21), dtype=bool)

        # Thicken laser lines (dilation) to about 7px
        if laser_mask.any():
            laser_thick = ndimage.binary_dilation(laser_mask, structure=_k20)
            out[laser_thick] = LASER_COLOR
        
        # Thicken v-groove lines (dilation) to about 7px
        if vg_mask.any():
            vg_thick = ndimage.binary_dilation(vg_mask, structure=_k20)
            out[vg_thick] = VGROOVE_COLOR
            
        # Thicken circle lines (dilation) to about 7px
        if circle_mask.any():
            circle_thick = ndimage.binary_dilation(circle_mask, structure=_k20)
            out[circle_thick] = CIRCLE_COLOR
            
        # Thicken bending lines (dilation) to about 7px
        if bending_mask.any():
            bending_thick = ndimage.binary_dilation(bending_mask, structure=_k20)
            out[bending_thick] = BENDING_COLOR

        # 6) Save as pure RGB (no alpha) to avoid print color shifts
        out_rgb = out.astype(np.uint8)
        out_img = Image.fromarray(out_rgb, mode="RGB")

        # 6b) Hard-monochrome for black-only parts to guarantee K-only output
        color_overlay_mask = ((out == CIRCLE_COLOR).all(axis=-1) |
                              (out == VGROOVE_COLOR).all(axis=-1) |
                              (out == BENDING_COLOR).all(axis=-1))
        if not color_overlay_mask.any():
            # Entire image is black-only → convert to true 1-bit bilevel (pure K-only)
            out_gray = np.dot(out_rgb[..., :3], [0.299, 0.587, 0.114]).astype(np.uint8)
            # Threshold at mid-level to avoid gray
            thresh = 128
            out_bw = (out_gray >= thresh).astype(np.uint8) * 255
            bw_img = Image.fromarray(out_bw, mode="L").convert("1")
            # Save bilevel PNG directly
            buf_bw = io.BytesIO()
            bw_img.save(buf_bw, format="PNG", optimize=True)
            return "data:image/png;base64," + base64.b64encode(buf_bw.getvalue()).decode("ascii")
        
        # 7) Resize back to fit within cell bounds (max 80x65px for part images)
        max_width, max_height = 80, 65
        if out_img.width > max_width or out_img.height > max_height:
            out_img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
        
        buf = io.BytesIO()
        # Force sRGB profile using a PIL encoder trick: save as JPEG (high quality, sRGB) then rewrap as PNG
        try:
            tmp = io.BytesIO()
            out_img.convert("RGB").save(tmp, format="JPEG", quality=95, subsampling=0)
            tmp.seek(0)
            from PIL import Image as _PILImage
            _PILImage.open(tmp).save(buf, format="PNG", optimize=True)
        except Exception:
            # Fallback to direct PNG save
            out_img.save(buf, format="PNG", optimize=True)
        return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")
        
    except Exception as e:
        print(f"Warning: Enhanced image recoloring failed: {e}")
        # Fallback to simple recoloring without thickening
        return _simple_recolor_png_data_url(data_url)

def _simple_recolor_png_data_url(data_url: str) -> str:
    """Simple recoloring without thickening (fallback function)."""
    if not _HAVE_PIL:
        return data_url
    
    try:
        m = re.match(r"^data:image/(png|jpeg);base64,(.+)$", data_url, flags=re.I|re.S)
        if not m:
            return data_url
        
        raw = base64.b64decode(m.group(2))
        im = Image.open(io.BytesIO(raw)).convert("RGBA")
        arr = np.array(im, dtype=np.uint8)
        rgb = arr[..., :3].astype(np.uint16)
        a = arr[..., 3:4]

        R, G, B = rgb[..., 0], rgb[..., 1], rgb[..., 2]

        # Simple masks
        m_bg = (R <= NEAR_BLACK_MAX) & (G <= NEAR_BLACK_MAX) & (B <= NEAR_BLACK_MAX)
        m_grid = (R <= GRID_MAX) & (G <= GRID_MAX) & (B <= GRID_MAX)
        m_vg = (R >= VG_R_MIN) & (G <= VG_G_MAX) & (B <= VG_B_MAX)
        m_laser = ((R >= LASER_MIN) & (G >= LASER_MIN) & (B >= LASER_MIN)) | \
                  ((R >= 100) & (G >= 100) & (B >= 100) & (R <= 200) & (G <= 200) & (B <= 200))
        m_circle = (R >= 140) & (R <= 200) & (G >= 100) & (G <= 160) & (B >= 100) & (B <= 160)
        m_bending = (R >= 120) & (R <= 180) & (G >= 120) & (G <= 180) & (B >= 180) & (B <= 250)
        m_alpha_bg = (a[..., 0] <= 10)

        out = rgb.copy()
        mask_bg_all = m_bg | m_grid | m_alpha_bg
        out[mask_bg_all] = BACKGROUND
        
        # Apply colors in priority order
        out[m_bending & ~mask_bg_all] = BENDING_COLOR
        out[m_circle & ~mask_bg_all & ~m_bending] = CIRCLE_COLOR
        out[m_vg & ~mask_bg_all & ~m_bending & ~m_circle] = VGROOVE_COLOR
        out[m_laser & ~mask_bg_all & ~m_bending & ~m_circle & ~m_vg] = LASER_COLOR

        out = np.concatenate([out.astype(np.uint8), np.full_like(a, 255)], axis=-1)
        out_img = Image.fromarray(out, mode="RGBA")
        buf = io.BytesIO()
        out_img.save(buf, format="PNG", optimize=True)
        return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")
        
    except Exception as e:
        print(f"Warning: Simple recoloring also failed: {e}")
        return data_url


# ----------------------- Pydantic Models (validation) -----------------------

class Company(BaseModel):
    name: str
    pscNumber: Optional[str] = None
    vatId: Optional[str] = None
    contactName: Optional[str] = None
    telephone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None

class PdfForm(BaseModel):
    quotationNumber: str
    quotationDate: str
    companyName: str
    pscNumber: Optional[str] = None
    vatId: Optional[str] = None
    contactName: Optional[str] = None
    telephone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    paymentTerms: Optional[str] = None
    deliveryMethod: Optional[str] = None
    deliveryPlace: Optional[str] = None
    deliveryDate: Optional[str] = None
    validity: Optional[str] = None

class MaterialInfo(BaseModel):
    material_name: str
    thickness: float
    grade: Optional[str] = None
    finish: Optional[str] = None
    scrap_factor: Optional[float] = None

class MaterialConfig(BaseModel):
    machine_speed: Optional[float] = None
    vaporization_speed: Optional[float] = None
    piercing_time: Optional[float] = None
    price_per_kg: Optional[float] = None
    density: Optional[float] = None
    vgroove_price: Optional[float] = None
    bending_price: Optional[float] = None

class AdminConfig(BaseModel):
    laser_cost: Optional[float] = None
    piercing_toggle: Optional[bool] = None

class Stats(BaseModel):
    original_entities: Optional[int] = None
    filtered_entities: Optional[int] = None
    removed_entities: Optional[int] = None
    layers_found: Optional[List[str]] = None
    total_parts: Optional[int] = None

class AreaCalcs(BaseModel):
    total_area: Optional[float] = None
    total_area_sq_mm: Optional[float] = None

class CalcDetails(BaseModel):
    weight_kg: Optional[float] = None
    volume_cm3: Optional[float] = None
    perimeter_meters: Optional[float] = None
    vgroove_count_total: Optional[int] = None
    vgroove_length_total: Optional[float] = None

class CostBreakdown(BaseModel):
    total_cost: Optional[float] = None
    laser_cost: Optional[float] = None
    material_cost: Optional[float] = None
    bending_cost: Optional[float] = None
    vgroove_cost: Optional[float] = None

class ServiceLine(BaseModel):
    enabled: Optional[bool] = None
    quantity: Optional[int] = None
    price: Optional[float] = None
    total: Optional[float] = None

class ExternalServices(BaseModel):
    punching: Optional[ServiceLine] = None
    brushing: Optional[ServiceLine] = None
    marking: Optional[ServiceLine] = None
    uv_print: Optional[ServiceLine] = None
    cutting: Optional[ServiceLine] = None
    pvc_cover: Optional[ServiceLine] = None
    rolling: Optional[ServiceLine] = None
    straighten: Optional[ServiceLine] = None
    cornerform: Optional[ServiceLine] = None
    router: Optional[ServiceLine] = None
    finishing: Optional[ServiceLine] = None
    installing: Optional[ServiceLine] = None
    other: Optional[ServiceLine] = None
    external_services_total: Optional[float] = None

class PartCostData(BaseModel):
    area_sq_mm: Optional[float] = None
    perimeter_meters: Optional[float] = None
    cutting_time_machine: Optional[float] = None
    cutting_time_vaporization: Optional[float] = None
    piercing_time_total: Optional[float] = None
    total_time_min: Optional[float] = None
    object_parts_count: Optional[int] = None
    laser_cost: Optional[float] = None
    weight_kg: Optional[float] = None
    material_cost: Optional[float] = None
    vgroove_count: Optional[int] = None
    bending_cost: Optional[float] = None
    vgroove_length_meters: Optional[float] = None
    vgroove_cost: Optional[float] = None
    total_cost: Optional[float] = None
    # Additional fields for the new table structure
    length_mm: Optional[float] = None
    width_mm: Optional[float] = None

class PartCost(BaseModel):
    part_number: Optional[int] = None
    area: Optional[float] = None
    object_parts_count: Optional[int] = None
    cost_data: Optional[PartCostData] = None

class PartDetails(BaseModel):
    part_costs: Optional[List[PartCost]] = None
    part_images: Optional[List[Dict[str, Any]]] = None

class Visualization(BaseModel):
    image: Optional[str] = None  # base64 PNG

class Payload(BaseModel):
    pdf_form_data: Optional[PdfForm] = None
    material_information: Optional[MaterialInfo] = None
    material_config: Optional[MaterialConfig] = None
    admin_config: Optional[AdminConfig] = None
    file_information: Optional[Dict[str, Any]] = None
    processing_statistics: Optional[Stats] = None
    area_calculations: Optional[AreaCalcs] = None
    calculation_details: Optional[CalcDetails] = None
    cost_breakdown: Optional[CostBreakdown] = None
    external_services_form_data: Optional[Dict[str, Any]] = None  # keep raw for table mapping
    part_details: Optional[PartDetails] = None
    visualization: Optional[Visualization] = None
    
    # Support for flat structure (Flask app format)
    quotationNumber: Optional[str] = None
    quotationDate: Optional[str] = None
    companyName: Optional[str] = None
    pscNumber: Optional[str] = None
    vatId: Optional[str] = None
    contactName: Optional[str] = None
    telephone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    paymentTerms: Optional[str] = None
    deliveryMethod: Optional[str] = None
    deliveryPlace: Optional[str] = None
    deliveryDate: Optional[str] = None
    validity: Optional[str] = None
    material_name: Optional[str] = None
    thickness: Optional[float] = None
    grade: Optional[str] = None
    finish: Optional[str] = None
    scrap_factor: Optional[float] = None
    
    # Rich data from Flask app processing
    stats: Optional[Dict[str, Any]] = None
    part_images: Optional[List[Dict[str, Any]]] = None
    part_costs: Optional[List[Dict[str, Any]]] = None
    image: Optional[str] = None  # Main visualization image
    
    # Additional fields that might be in the payload
    original_entities: Optional[int] = None
    filtered_entities: Optional[int] = None
    removed_entities: Optional[int] = None
    layers_found: Optional[List[str]] = None
    total_parts: Optional[int] = None
    total_area: Optional[float] = None
    total_cost: Optional[float] = None
    filename: Optional[str] = None
    original_filename: Optional[str] = None
    
    class Config:
        extra = "allow"  # Allow additional fields


# ----------------------- Renderer Classes -----------------------

class PlaywrightChromiumRenderer:
    """Primary renderer using Playwright + Chromium headless."""
    
    @staticmethod
    def ensure_playwright_installed():
        """Try to install chromium if not available."""
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch()
                browser.close()
                return True
        except Exception:
            try:
                print("Installing Playwright Chromium...")
                subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"], check=True)
                return True
            except Exception as e:
                raise RuntimeError(
                    f"Playwright Chromium not available. Please run: python -m playwright install chromium\n"
                    f"Error: {e}"
                )
    
    @classmethod
    def print_pdf(cls, html_string: str, base_url: str, out_pdf: str, html_title: str = "Document") -> str:
        """Generate PDF using Playwright + Chromium."""
        cls.ensure_playwright_installed()
        
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.set_content(html_string, wait_until="networkidle")
            
            Path(out_pdf).parent.mkdir(parents=True, exist_ok=True)
            page.pdf(
                path=out_pdf,
                format="A4",
                margin={
                    "top": "18mm",
                    "right": "16mm", 
                    "bottom": "20mm",
                    "left": "16mm"
                },
                display_header_footer=True,
                header_template="<div></div>",
                footer_template='<div style="font-size:10px; text-align:center; width:100%;">Page <span class="pageNumber"></span> of <span class="totalPages"></span></div>',
                print_background=True
            )
            browser.close()
        
        return out_pdf


class WeasyPrintRenderer:
    """Fallback renderer using WeasyPrint."""
    
    @classmethod  
    def write_pdf(cls, html_string: str, base_url: str, out_pdf: str) -> str:
        """Generate PDF using WeasyPrint."""
        if not _HAVE_WEASY:
            raise RuntimeError("WeasyPrint not available. Install with: pip install weasyprint")
        
        Path(out_pdf).parent.mkdir(parents=True, exist_ok=True)
        HTML(string=html_string, base_url=base_url).write_pdf(out_pdf)
        return out_pdf


class WkhtmltopdfRenderer:
    """Fallback renderer using wkhtmltopdf binary."""
    
    @classmethod
    def write_pdf(cls, html_string: str, base_url: str, out_pdf: str) -> str:
        """Generate PDF using wkhtmltopdf."""
        if not _HAVE_WKHTMLTOPDF:
            raise RuntimeError("wkhtmltopdf not available. Install wkhtmltopdf and ensure it's in PATH.")
        
        import tempfile
        
        Path(out_pdf).parent.mkdir(parents=True, exist_ok=True)
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False) as f:
            f.write(html_string)
            html_file = f.name
        
        try:
            subprocess.run([
                "wkhtmltopdf",
                "--page-size", "A4",
                "--margin-top", "18mm",
                "--margin-right", "16mm", 
                "--margin-bottom", "20mm",
                "--margin-left", "16mm",
                "--footer-center", "Page [page] of [topage]",
                "--footer-font-size", "10",
                html_file,
                out_pdf
            ], check=True)
        finally:
            Path(html_file).unlink(missing_ok=True)
        
        return out_pdf


# ----------------------- HTML Template & CSS -----------------------

HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{{ title }}</title>
<style>
{{ css }}
</style>
</head>
<body>

<!-- Header with Logo and Quotation Number -->
<header class="quotation-header">
  {% if logo_url %}
  <div class="logo-container">
    <img class="logo" src="{{ logo_url }}" alt="Company Logo">
  </div>
  {% endif %}
  <div class="quotation-date">Date: {{ format_date_dd_mm_yyyy(pdf.quotationDate) or "–" }}</div>
</header>
<div class="quotation-info">
<h1>Quotation Number: {{ pdf.quotationNumber or "–" }}</h1>
    <h3>Budget Price +/-5%</h3>
  </div>
  <hr>
<!-- Opening Statement -->
<div class="opening-statement">
  <p><strong>Mr/Mrs {{ pdf.contactName or "Customer" }},</strong></p>
  <p>We thank you for your inquiry and we are pleased to give you a fast budget price for your order.</p>
  </div>

<!-- Company Information Block -->
<section class="company-block">
  <div class="company-info">
    <div class="company-name"><strong>Company Name:</strong> {{ pdf.companyName or "–" }}</div>
    <div class="company-details">
      <div><strong>PSC Number:</strong> {{ pdf.pscNumber or "–" }}</div>
      <div><strong>VAT ID:</strong> {{ pdf.vatId or "–" }}</div>
      <div><strong>Contact Name:</strong> {{ pdf.contactName or "–" }}</div>
      <div><strong>Telephone:</strong> {{ pdf.telephone or "–" }}</div>
      <div><strong>Email:</strong> {{ pdf.email or "–" }}</div>
  </div>
    <div class="company-address">
      <strong>Address:</strong><br>
      {{ pdf.address or "–" }}
  </div>
  </div>
</section>

<!-- Items Table -->
<section class="items-section">
  <h2>Items</h2>
  <table class="items-table">
    <thead>
      <tr>
        <th>Part Image</th>
        <th>Material</th>
        <th>Grade/Finish</th>
        <th>Quantity</th>
        <th>Length (mm)</th>
        <th>Width (mm)</th>
        <th>Thickness (mm)</th>
        <th>Material Price</th>
        <th>Laser Price</th>
        <th>Bending Price</th>
        <th>V-Groove Price</th>
        <th>Sub Total $</th>
      </tr>
    </thead>
    <tbody>
      {% for p in parts %}
      <tr>
        <td class="part-image-cell">
  {% if part_images %}
    {% for part_img in part_images if part_img.part_number == p.part_number %}
      {% if part_img.svg %}
        {# ── SVG branch: force a box so Chrome prints it ── #}
        <img class="part-thumbnail part-thumbnail--mono"
             src="data:image/svg+xml;base64,{{ part_img.svg }}"
             alt="Part {{ p.part_number }}"
             width="80" height="65">
      {% else %}
        {# Use PNG directly (with recoloring) to preserve curves like SPLINE/ELLIPSE #}
        <img class="part-thumbnail part-thumbnail--mono"
             src="{{ recolor_image('data:image/png;base64,' + part_img.image) }}"
             alt="Part {{ p.part_number }}">
      {% endif %}
    {% endfor %}
  {% else %}
    <div class="no-image">–</div>
  {% endif %}
</td>

        {# Resolve per-part material from associated data if available #}
        {% set part_meta = None %}
        {% if part_images %}
          {% for part_img in part_images if part_img.part_number == p.part_number %}
            {% set part_meta = part_img %}
          {% endfor %}
        {% endif %}
        {% if not part_meta and p is mapping %}
          {% set part_meta = p %}
        {% endif %}
        {# Material per-part only: applied_* then raw; if none, show 0 #}
        {% set mat_name =
             (part_meta.applied_material_name if part_meta and part_meta.applied_material_name
              else (part_meta.material_name if part_meta and part_meta.material_name else None)) %}
        <td>{{ mat_name if mat_name else '0' }}</td>
        <td>
          {# If material is NOT Stainless Steel, force 0/0 #}
          {% set mat_lower = (mat_name|string).lower() if mat_name else '' %}
          {% if 'stainless' not in mat_lower %}
            0 / 0
          {% else %}
            {% set g = (part_meta.applied_grade if part_meta and part_meta.applied_grade is not none and part_meta.applied_grade|string|length > 0
                        else (part_meta.grade if part_meta and part_meta.grade is not none and part_meta.grade|string|length > 0 else '0')) %}
            {% set f = (part_meta.applied_finish if part_meta and part_meta.applied_finish is not none and part_meta.applied_finish|string|length > 0
                        else (part_meta.finish if part_meta and part_meta.finish is not none and part_meta.finish|string|length > 0 else '0')) %}
            {{ g }} / {{ f }}
          {% endif %}
        </td>
        <td class="num">{{ p.object_parts_count or (p.cost_data.object_parts_count if p.cost_data else '–') }}</td>
        <td class="num">{{ format_number(p.cost_data.length_mm, 2) if p.cost_data and p.cost_data.length_mm else '–' }}</td>
        <td class="num">{{ format_number(p.cost_data.width_mm, 2) if p.cost_data and p.cost_data.width_mm else '–' }}</td>
        {# Thickness per-part only; missing => 0 #}
        {% set thk = (part_meta.applied_thickness if part_meta and part_meta.applied_thickness is not none
                       else (part_meta.thickness if part_meta and part_meta.thickness is not none else None)) %}
        <td class="num">{{ thk if thk is not none else '0' }}</td>
        <td class="num">{{ format_currency(p.cost_data.material_cost) if p.cost_data and p.cost_data.material_cost else '–' }}</td>
        <td class="num">{{ format_currency(p.cost_data.laser_cost) if p.cost_data and p.cost_data.laser_cost else '–' }}</td>
        <td class="num">{{ format_currency(p.cost_data.bending_cost) if p.cost_data and p.cost_data.bending_cost else '$0.00' }}</td>
        <td class="num">{{ format_currency(p.cost_data.vgroove_cost) if p.cost_data and p.cost_data.vgroove_cost else '$0.00' }}</td>
        <td class="num total-cell">{{ format_currency(p.cost_data.total_cost) if p.cost_data and p.cost_data.total_cost else '–' }}</td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
</section>



<!-- Totals Section -->
<section class="totals-section">
  <div class="totals-list">

    {# -------- Laser price -------- #}
    {% set laser_price = costs.get('total_cost') or costs.get('laser_cost') or 0 %}
    <div class="total-item">
      <span class="total-label">Laser Price:</span>
      <span class="total-value">{{ format_currency(laser_price) }}</span>
    </div>

    {# ----------- External services (Add …) ------------ #}
    {# In Jinja2 a plain variable set inside a loop is lost outside that loop,
       so we use a “namespace” object that survives the scope. #}
    {% set ns = namespace(ext_total=0) %}

    {% for name, line in ext_lines %}
      {% if line.enabled and line.total and line.total > 0 %}
        {% set ns.ext_total = ns.ext_total + line.total %}
        <div class="total-item">
          <span class="total-label">Add {{ name.replace('_', ' ').title() }}:</span>
          <span class="total-value">{{ format_currency(line.total) }}</span>
        </div>
      {% endif %}
    {% endfor %}

    {# -------- Sub-total, VAT, Grand total -------- #}
    {% set subtotal     = laser_price + ns.ext_total %}
    {% set vat_amount   = subtotal * 0.11 %}
    {% set final_total  = subtotal + vat_amount %}

    <div class="total-item">
      <span class="total-label">SubTotal VAT Excluded:</span>
      <span class="total-value">{{ format_currency(subtotal) }}</span>
    </div>

    <div class="total-item">
      <span class="total-label">VAT 11%:</span>
      <span class="total-value">{{ format_currency(vat_amount) }}</span>
    </div>

    <div class="total-item total-final">
      <span class="total-label">Total VAT Included:</span>
      <span class="total-value">{{ format_currency(final_total) }}</span>
    </div>

  </div>
</section>

<!-- Delivery Information -->
<section class="delivery-section">
  <div class="delivery-grid">
    <div class="delivery-item">
      <strong>Delivery Place:</strong> {{ pdf.deliveryPlace or "–" }}
    </div>
    <div class="delivery-item">
      <strong>Delivery Date:</strong> {{ pdf.deliveryDate or "–" }}
    </div>
    <div class="delivery-item">
      <strong>Delivery Method:</strong> {{ pdf.deliveryMethod or "–" }}
    </div>
    <div class="delivery-item">
      <strong>Payment Terms:</strong> {{ pdf.paymentTerms or "Standard payment terms apply" }}
    </div>
  </div>
</section>

<!-- Prepared By Section -->
<section class="prepared-by-section">
  <div class="prepared-by">
    <strong>Quotation prepared By:</strong> {{ prepared_by or pdf.contactName or "–" }}
    <div class="extension">Ext: {{ extension_number if extension_number and extension_number != '' else (pdf.telephone if pdf.telephone else "–") }}</div>
  </div>
</section>

<!-- Sales Terms & Conditions -->
<section class="terms-section">
  <h3>Sales Terms & Conditions:</h3>
  <p>All prices are subject to change without prior notice. Payment terms are as agreed upon between parties.</p>
  
  <h3>Validity: {{ pdf.validity or "–" }}</h3>
  <p>N.B.: In the event of any claim on the goods received, it should be notified to us within 24 hours after reception of your order, 
  any item left over after your pickup will be considered as scrap. Please note that your order can be processed only in case all your previous 
  invoices have been settled on basis of our agreement terms. We hope our offer is satisfactory and look forward to receive your confirmation to
   which we shall give our prompt and careful attention. Should you need any further information, please feel free to contact us.</p>
  
  <div class="company-signature">
    <strong>Naggiar Trading SAL</strong><br>
    <em>Please contact us for any questions regarding this quotation.</em>
  </div>
</section>

</body>
</html>
"""

CSS_DEFAULT = r"""
@page {
  size: A4; 
  margin: 18mm 16mm 20mm 16mm; 
}

body { 
  font-family: Helvetica, Arial, sans-serif; 
  font-size: 11px; 
  line-height: 1.3;
  color: #000;
  margin: 0;
  padding: 0;
}

/* Header Section */
.quotation-header {
  display: flex;
  justify-content: space-between;
  align-items: flex-start;
  margin-bottom: 20px;
  padding-bottom: 10px;
  border-bottom: 2px solid #000;
  position: relative;
}

.quotation-info h1 {
  font-size: 18px;
  font-weight: bold;
  margin: 0 0 5px 0;
  color: #000;
}

.quotation-date {
  font-size: 12px;
  color: #333;
  position: absolute;
  top: 0;
  right: 0;
  font-weight: bold;
}

.logo-container {
  text-align: right;
}

.logo {
  height: 150px;
  width: 100%;
  object-fit: contain;
}

/* Opening Statement */
.opening-statement {
  margin-bottom: 20px;
}

.opening-statement p {
  margin: 5px 0;
  font-size: 11px;
}

/* Company Block */
.company-block {
  margin-bottom: 20px;
  padding: 10px;
  border: 1px solid #ccc;
  background: #f9f9f9;
}

.company-name {
  font-size: 14px;
  font-weight: bold;
  margin-bottom: 8px;
  color: #000;
}

.company-details {
  margin-bottom: 8px;
}

.company-details div {
  margin-bottom: 3px;
  font-size: 10px;
}

.company-address {
  font-size: 10px;
  white-space: pre-line;
  line-height: 1.2;
}

/* Items Table */
.items-section {
  margin-bottom: 20px;
}

.items-section h2 {
  font-size: 14px;
  font-weight: bold;
  margin: 0 0 10px 0;
  color: #000;
  border-bottom: 1px solid #000;
  padding-bottom: 3px;
}

.items-table {
  width: 100%;
  border-collapse: collapse;
  font-size: 9px;
  page-break-inside: auto;
}

.items-table th,
.items-table td {
  border: 1px solid #000;
  padding: 4px 6px;
  vertical-align: middle;
  text-align: center;
}

/* Subtotal column styling */
.items-table th:last-child,
.items-table td:last-child {
  background-color: #808080;
  color: white;
  font-weight: bold;
}

.items-table th {
  background: #f0f0f0;
  font-weight: bold;
  font-size: 8px;
}

.items-table .num {
  text-align: center;
  font-family: "Courier New", monospace;
}

.items-table .total-cell {
  font-weight: bold;
  background: #808080;
  color: white;
  text-align: center;
}

.part-image-cell {
  width: 80px;
  text-align: center;
  vertical-align: middle;
}

.part-thumbnail {
  max-width: 80px;
  max-height: 65px;
  object-fit: contain;
  border: none;
  background: white;
  padding: 0;
  border-radius: 0;
  margin: 0;
}

/* Force printer-safe rendering for black-only content */
.part-thumbnail--mono {
  image-rendering: crisp-edges;
  filter: grayscale(100%); /* grayscale in compositor → prevents CMY tint for black parts */
  color-interpolation-filters: sRGB;
}

.no-image {
  color: #999;
  font-style: italic;
}

/* External Services */
.external-services-section {
  margin-bottom: 20px;
}

.external-services-section h2 {
  font-size: 14px;
  font-weight: bold;
  margin: 0 0 10px 0;
  color: #000;
  border-bottom: 1px solid #000;
  padding-bottom: 3px;
}

.addons-list {
  margin-left: 20px;
}

.addon-item {
  display: flex;
  justify-content: space-between;
  margin-bottom: 5px;
  font-size: 11px;
}

.addon-label {
  font-weight: bold;
}

.addon-price {
  font-family: "Courier New", monospace;
  font-weight: bold;
}

/* Totals Section */
.totals-section {
  margin-bottom: 20px;
  padding: 15px;
  background: #ffffff;
  border: 1px solid #ddd;
  border-radius: 4px;
}

.totals-list {
  display: flex;
  flex-direction: column;
  gap: 8px;
}

.total-item {
  display: flex;
  justify-content: space-between;
  align-items: center;
  font-size: 11px;
  font-family: Helvetica, Arial, sans-serif;
  padding: 4px 0;
}

.total-label {
  font-weight: normal;
  color: #333;
}

.total-value {
  font-weight: normal;
  color: #333;
  text-align: right;
}

.total-final {
  font-weight: bold;
  font-size: 12px;
  border-top: 1px solid #ccc;
  padding-top: 8px;
  margin-top: 4px;
}

/* Delivery Section */
.delivery-section {
  margin-bottom: 20px;
}

.delivery-grid {
  display: grid;
  grid-template-columns: repeat(2, 1fr);
  gap: 10px;
  font-size: 10px;
}

.delivery-item {
  padding: 5px;
  border: 1px solid #ccc;
  background: #f9f9f9;
}

/* Prepared By Section */
.prepared-by-section {
  margin-bottom: 20px;
  text-align: right;
}

.prepared-by {
  font-size: 11px;
  font-weight: bold;
}

.extension {
  font-size: 10px;
  color: #666;
  margin-top: 3px;
}

/* Terms Section */
.terms-section {
  margin-bottom: 20px;
  font-size: 10px;
  line-height: 1.4;
}

.terms-section h3 {
  font-size: 12px;
  font-weight: bold;
  margin: 10px 0 5px 0;
  color: #000;
}

.terms-section p {
  margin: 5px 0;
  text-align: justify;
}

.company-signature {
  margin-top: 15px;
  text-align: center;
  font-size: 11px;
}

.company-signature strong {
  font-size: 12px;
}

.company-signature em {
  font-size: 10px;
  color: #666;
}

/* Print optimizations */
@media print {
  .items-table thead {
    display: table-header-group;
  }
  
  .items-table tfoot {
    display: table-footer-group;
  }
  
  tr {
    page-break-inside: avoid;
  }
  
  .items-table {
    page-break-inside: auto;
  }
  
  .totals-section {
    page-break-inside: avoid;
  }
  
  .terms-section {
    page-break-inside: avoid;
  }
}

/* Responsive adjustments */
@media screen {
  .items-table {
    font-size: 8px;
  }
  
  .items-table th {
    font-size: 7px;
  }
}
"""


# ----------------------- Helper Functions -----------------------

def format_number(value, decimals: int = 2):
    """Format a number with specified decimals, return empty string if None/invalid."""
    if value is None:
        return ""
    try:
        return f"{float(value):.{decimals}f}"
    except (ValueError, TypeError):
        return ""

def format_currency(value, symbol: str = "$"):
    """Format a currency value with 2 decimals."""
    if value is None:
        return ""
    try:
        return f"{symbol}{float(value):.2f}"
    except (ValueError, TypeError):
        return ""

def format_thickness(value):
    """Format thickness with units."""
    if value is None:
        return "–"
    try:
        return f"{float(value):.1f} mm"
    except (ValueError, TypeError):
        return "–"

def format_meters(value):
    """Format meters with units."""
    if value is None:
        return "–"
    try:
        return f"{float(value):.3f} m"
    except (ValueError, TypeError):
        return "–"

def format_weight(value):
    """Format weight with units."""
    if value is None:
        return "–"
    try:
        return f"{float(value):.3f} kg"
    except (ValueError, TypeError):
        return "–"

def format_area(area_m2, area_mm2):
    """Format area with both m² and mm² units."""
    if area_m2 is None and area_mm2 is None:
        return "–"
    
    m2_str = f"{float(area_m2):.6f} m²" if area_m2 is not None else "– m²"
    mm2_str = f"{float(area_mm2):.0f} mm²" if area_mm2 is not None else "– mm²"
    
    return f"{m2_str} ({mm2_str})"

def format_vgroove(count, length):
    """Format V-groove count and length."""
    if count is None and length is None:
        return "–"
    
    count_str = str(int(count)) if count is not None else "–"
    length_str = f"{float(length):.3f} m" if length is not None else "– m"
    
    return f"{count_str} lines / {length_str}"

def format_list(items):
    """Format a list as comma-separated string."""
    if not items:
        return ""
    return ", ".join(str(item) for item in items)

def format_date_dd_mm_yyyy(date_str):
    """Format date string to DD-MM-YYYY format."""
    if not date_str:
        return ""
    try:
        # Try to parse various date formats
        from datetime import datetime
        # Common formats to try
        formats = [
            "%Y-%m-%d",      # 2025-01-15
            "%d/%m/%Y",      # 15/01/2025
            "%m/%d/%Y",      # 01/15/2025
            "%d-%m-%Y",      # 15-01-2025
            "%Y-%m-%d %H:%M:%S",  # 2025-01-15 10:30:00
        ]
        
        parsed_date = None
        for fmt in formats:
            try:
                parsed_date = datetime.strptime(date_str, fmt)
                break
            except ValueError:
                continue
        
        if parsed_date:
            return parsed_date.strftime("%d-%m-%Y")
        else:
            # If we can't parse it, return as-is
            return date_str
    except Exception:
        return date_str


# ----------------------- Core Service -----------------------

class DocGenerator:
    def __init__(self, *, excel_template: Optional[str] = None, cell_map: Optional[Dict[str, str]] = None):
        self.excel_template = Path(excel_template) if excel_template else None
        self.cell_map = cell_map or {}

    # -------- Excel filling (unchanged) --------
    def _get(self, data: Dict[str, Any], dotted: str) -> Any:
        """Retrieve nested dotted path value."""
        cur = data
        for part in dotted.split("."):
            if "[" in part and part.endswith("]"):
                # array index
                key, idx = part[:-1].split("[")
                cur = cur.get(key, [])
                cur = cur[int(idx)] if len(cur) > int(idx) else None
            else:
                if not isinstance(cur, dict): return None
                cur = cur.get(part)
            if cur is None:
                return None
        return cur

    def _insert_excel_image(self, ws, anchor_cell: str, b64_png: str):
        from io import BytesIO
        try:
            img_bytes = base64.b64decode(b64_png)
        except Exception:
            return
        bio = BytesIO(img_bytes)
        try:
            img = XLImage(bio)
            ws.add_image(img, anchor_cell)
        except Exception:
            pass

    def fill_excel(self, payload: Dict[str, Any], out_xlsx: str, *, image_key: Optional[str] = "visualization.image") -> Optional[str]:
        if not self.excel_template:
            return None
        wb = load_workbook(self.excel_template)
        ws = wb["Quote"] if "Quote" in wb.sheetnames else wb.active

        # write mapped values
        for key, cell in self.cell_map.items():
            if key.endswith("_anchor"):
                continue
            val = self._get(payload, key)
            if val is None:
                continue
            
            # Skip merged cells
            cell_obj = ws[cell]
            if hasattr(cell_obj, 'coordinate') and any(cell_obj.coordinate in merged_range for merged_range in ws.merged_cells.ranges):
                print(f"Warning: Skipping merged cell {cell} for key {key}")
                continue
            
            ws[cell] = ", ".join(val) if isinstance(val, list) else val

        # optional image insertion
        if image_key:
            anchor = self.cell_map.get(f"{image_key}_anchor") or self.cell_map.get("visualization.image_anchor")
            img_b64 = self._get(payload, image_key)
            if anchor and img_b64:
                # Recolor for clarity (same as PDF) and enforce pure RGB
                try:
                    data_url = recolor_png_data_url(f"data:image/png;base64,{img_b64}")
                    m = re.match(r"^data:image/(png|jpeg);base64,(.+)$", data_url, flags=re.I|re.S)
                    if m:
                        img_b64 = m.group(2)
                except Exception:
                    pass
                self._insert_excel_image(ws, anchor, img_b64)

        Path(out_xlsx).parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_xlsx)
        return out_xlsx

    # -------- PDF rendering with pluggable renderers --------
    def render_pdf(self, payload: Dict[str, Any], out_pdf: str, *, title: str = "Quotation",
                   logo_path: Optional[str] = None, css_override: Optional[str] = None,
                   template_override: Optional[str] = None, base_url: Optional[str] = None) -> str:
        """Render PDF using pluggable renderer architecture."""

        # Validate payload
        data = Payload(**payload)

        # Prepare template context
        css_str = Path(css_override).read_text() if css_override else CSS_DEFAULT
        html_tpl = Path(template_override).read_text() if template_override else HTML_TEMPLATE

        env = Environment(loader=BaseLoader(), autoescape=select_autoescape())
        
        # Add formatting functions to template environment
        env.globals.update({
            'format_number': format_number,
            'format_currency': format_currency,
            'format_thickness': format_thickness,
            'format_meters': format_meters,
            'format_weight': format_weight,
            'format_area': format_area,
            'format_vgroove': format_vgroove,
            'format_list': format_list,
            'format_date_dd_mm_yyyy': format_date_dd_mm_yyyy,
        })
        
        tmpl = env.from_string(html_tpl)

        # Prepare external services data
        ext_lines = []
        es = data.external_services_form_data or {}
        svc_names = ["punching","brushing","marking","uv_print","cutting","pvc_cover","rolling","straighten","cornerform","router","finishing","installing","other"]
        
        for s in svc_names:
            # Handle both data formats:
            # 1. New format: {service}_enabled, {service}_quantity, etc.
            # 2. Frontend format: {service: {quantity, price, total}}
            
            if f"{s}_enabled" in es:
                # Format 1: Individual fields
                line = ServiceLine(
                    enabled=es.get(f"{s}_enabled"),
                    quantity=es.get(f"{s}_quantity"),
                    price=es.get(f"{s}_price"),
                    total=es.get(f"{s}_total"),
                )
            elif s in es:
                # Format 2: Service object
                service_data = es[s]
                line = ServiceLine(
                    enabled=True,  # If it's in the data, it's enabled
                    quantity=service_data.get("quantity", 0),
                    price=service_data.get("price", 0),
                    total=service_data.get("total", 0),
                )
            else:
                # Service not present
                line = ServiceLine(
                    enabled=False,
                    quantity=0,
                    price=0,
                    total=0,
                )
            
            ext_lines.append((s, line))

        # Helper function to get data from either nested or flat structure
        def get_nested_or_flat(nested_obj, flat_prefix):
            if nested_obj:
                return nested_obj.model_dump()
            # Build from flat structure
            result = {}
            for key, value in data.model_dump().items():
                if key.startswith(flat_prefix) or key in ['quotationNumber', 'quotationDate', 'companyName', 'pscNumber', 'vatId', 'contactName', 'telephone', 'email', 'address', 'paymentTerms', 'deliveryMethod', 'deliveryPlace', 'deliveryDate', 'validity', 'material_name', 'thickness', 'grade', 'finish', 'scrap_factor']:
                    if flat_prefix == 'pdf' and key in ['quotationNumber', 'quotationDate', 'companyName', 'pscNumber', 'vatId', 'contactName', 'telephone', 'email', 'address', 'paymentTerms', 'deliveryMethod', 'deliveryPlace', 'deliveryDate', 'validity']:
                        result[key] = value
                    elif flat_prefix == 'mat' and key in ['material_name', 'thickness', 'grade', 'finish', 'scrap_factor']:
                        result[key] = value
            return result

        # Helper function to extract data from the rich payload structure
        def extract_rich_data():
            # Get data from the payload
            payload_dict = data.model_dump()
            
            # Extract stats (processing results)
            stats = payload_dict.get('stats', {})
            if not stats and 'original_entities' in payload_dict:
                # Fallback: build stats from flat structure
                stats = {
                    'original_entities': payload_dict.get('original_entities'),
                    'filtered_entities': payload_dict.get('filtered_entities'),
                    'removed_entities': payload_dict.get('removed_entities'),
                    'layers_found': payload_dict.get('layers_found', []),
                    'total_parts': payload_dict.get('total_parts'),
                    'filename': payload_dict.get('filename'),
                    'total_area': payload_dict.get('total_area'),
                    'total_cost': payload_dict.get('total_cost')
                }
            
            # Extract part details and enhance with dimensions
            part_details = payload_dict.get('part_details', {})
            if not part_details:
                part_costs = payload_dict.get('part_costs', [])
                # Enhance part costs with actual dimensions from frontend
                for part_cost in part_costs:
                    if isinstance(part_cost, dict) and 'cost_data' in part_cost:
                        cost_data = part_cost['cost_data']
                        if isinstance(cost_data, dict):
                            # Use actual length and width from the part data
                            if 'length_mm' in part_cost and part_cost['length_mm'] is not None:
                                cost_data['length_mm'] = part_cost['length_mm']
                            if 'width_mm' in part_cost and part_cost['width_mm'] is not None:
                                cost_data['width_mm'] = part_cost['width_mm']
                            
                            # Only calculate from area if length/width are not available
                            if 'length_mm' not in cost_data or 'width_mm' not in cost_data:
                                area_sq_mm = cost_data.get('area_sq_mm') or (cost_data.get('area', 0) * 1000000)
                                if area_sq_mm and area_sq_mm > 0:
                                    # Assume square-ish parts, calculate length and width
                                    import math
                                    side_length = math.sqrt(area_sq_mm)
                                    if 'length_mm' not in cost_data:
                                        cost_data['length_mm'] = round(side_length, 2)
                                    if 'width_mm' not in cost_data:
                                        cost_data['width_mm'] = round(side_length, 2)
                
                part_details = {
                    'part_costs': part_costs,
                    'part_images': payload_dict.get('part_images', [])
                }
            
            # Extract material config
            material_config = payload_dict.get('material_config', {})
            if not material_config:
                # Build from flat structure
                material_config = {
                    'machine_speed': payload_dict.get('machine_speed'),
                    'vaporization_speed': payload_dict.get('vaporization_speed'),
                    'piercing_time': payload_dict.get('piercing_time'),
                    'price_per_kg': payload_dict.get('price_per_kg'),
                    'density': payload_dict.get('density'),
                    'vgroove_price': payload_dict.get('vgroove_price'),
                    'bending_price': payload_dict.get('bending_price')
                }
            
            # Extract cost breakdown
            cost_breakdown = payload_dict.get('cost_breakdown', {})
            if not cost_breakdown:
                # Calculate totals from part costs
                total_bending = 0
                total_vgroove = 0
                total_material = 0
                total_laser = 0
                total_cost = 0
                
                part_costs = payload_dict.get('part_costs', [])
                for part in part_costs:
                    if isinstance(part, dict) and 'cost_data' in part:
                        cost_data = part['cost_data']
                        if isinstance(cost_data, dict):
                            total_bending += cost_data.get('bending_cost', 0) or 0
                            total_vgroove += cost_data.get('vgroove_cost', 0) or 0
                            total_material += cost_data.get('material_cost', 0) or 0
                            total_laser += cost_data.get('laser_cost', 0) or 0
                            total_cost += cost_data.get('total_cost', 0) or 0
                
                cost_breakdown = {
                    'total_cost': total_cost or stats.get('total_cost'),
                    'laser_cost': total_laser or payload_dict.get('laser_cost'),
                    'material_cost': total_material or payload_dict.get('material_cost'),
                    'bending_cost': total_bending or payload_dict.get('bending_cost'),
                    'vgroove_cost': total_vgroove or payload_dict.get('vgroove_cost')
                }
            
            # Extract area calculations
            area_calculations = payload_dict.get('area_calculations', {})
            if not area_calculations:
                area_calculations = {
                    'total_area': stats.get('total_area'),
                    'total_area_sq_mm': payload_dict.get('total_area_sq_mm')
                }
            
            # Extract calculation details
            calculation_details = payload_dict.get('calculation_details', {})
            if not calculation_details:
                calculation_details = {
                    'weight_kg': payload_dict.get('weight_kg'),
                    'volume_cm3': payload_dict.get('volume_cm3'),
                    'perimeter_meters': payload_dict.get('perimeter_meters'),
                    'vgroove_count_total': payload_dict.get('vgroove_count_total'),
                    'vgroove_length_total': payload_dict.get('vgroove_length_total')
                }
            
            return {
                'stats': stats,
                'part_details': part_details,
                'material_config': material_config,
                'cost_breakdown': cost_breakdown,
                'area_calculations': area_calculations,
                'calculation_details': calculation_details
            }

        # Helper function to safely get nested object or empty dict
        def safe_model_dump(obj):
            if obj:
                return obj.model_dump()
            return {}

        # Extract rich data from the payload
        rich_data = extract_rich_data()
        
        # Read logo as base64 data URI
        logo_data_uri = None
        if logo_path and Path(logo_path).exists():
            with open(logo_path, "rb") as f:
                logo_bytes = f.read()
                logo_b64 = base64.b64encode(logo_bytes).decode("utf-8")
                ext = Path(logo_path).suffix.lower().replace('.', '')
                logo_data_uri = f"data:image/{ext};base64,{logo_b64}"

        # Helper to get part field from part, then cfg, then mat, then payload
        def get_part_field(part, field, cfg, mat, payload):
            # Try part
            if part and part.get(field):
                return part.get(field)
            # Try cost_data
            if part and part.get('cost_data') and part['cost_data'].get(field):
                return part['cost_data'][field]
            # Try cfg
            if cfg and (cfg.get(field) or cfg.get(field.title()) or cfg.get(field.upper())):
                return cfg.get(field) or cfg.get(field.title()) or cfg.get(field.upper())
            # Try mat
            if mat and mat.get(field):
                return mat.get(field)
            # Try payload
            if payload and payload.get(field):
                return payload.get(field)
            return None

        # Template context with rich data support
        extension_number = getattr(data, 'extension_number', None) or ''
        print(f"DEBUG: Template context - extension_number: '{extension_number}'")
        
        ctx = {
            "title": title,
            "now": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "logo_url": logo_data_uri,
            "pdf": get_nested_or_flat(data.pdf_form_data, 'pdf'),
            "mat": {
                'material_name': getattr(data, 'material_name', None) or 'Stainless Steel',
                'thickness': getattr(data, 'thickness', None) or 2.0,
                'grade': getattr(data, 'grade', None) or '304',
                'finish': getattr(data, 'finish', None) or 'Brushed',
                'scrap_factor': getattr(data, 'scrap_factor', None) or 1.20
            },
            "cfg": rich_data['material_config'],
            "admin": safe_model_dump(data.admin_config),
            "file": rich_data['stats'],
            "stats": rich_data['stats'],
            "area": rich_data['area_calculations'],
            "calc": rich_data['calculation_details'],
            "costs": rich_data['cost_breakdown'],
            "ext": es,
            "ext_lines": ext_lines,
            "parts": rich_data['part_details'].get('part_costs', []),
            "part_images": rich_data['part_details'].get('part_images', []),
            "main_image": recolor_png_data_url(f"data:image/png;base64,{data.visualization.image}") if data.visualization and data.visualization.image else None,
            "css": css_str,
            "prepared_by": getattr(data, 'prepared_by', None) or getattr(data, 'contactName', None),
            "extension_number": getattr(data, 'extension_number', None),
            "get_part_field": get_part_field,
            "payload": data.model_dump() if hasattr(data, 'model_dump') else data,
            # DIRECT MATERIAL DATA - GUARANTEED TO WORK
            "material_name": getattr(data, 'material_name', None) or 'Stainless Steel',
            "thickness": getattr(data, 'thickness', None) or 2.0,
            "grade": getattr(data, 'grade', None) or '304',
            "finish": getattr(data, 'finish', None) or 'Brushed',
            # IMAGE RECOLORING FUNCTION
            "recolor_image": recolor_png_data_url,
            # SVG CONVERSION FUNCTION
            "convert_to_svg": png_to_svg_data_url,
        }

        html = tmpl.render(**ctx)
        base_url_resolved = base_url or str(Path.cwd())

        # Try renderers in order of preference
        errors = []
        
        # 1. Try Playwright (primary)
        if _HAVE_PLAYWRIGHT:
            try:
                return PlaywrightChromiumRenderer.print_pdf(html, base_url_resolved, out_pdf, title)
            except Exception as e:
                errors.append(f"Playwright: {e}")
        
        # 2. Try WeasyPrint (fallback)
        if _HAVE_WEASY:
            try:
                return WeasyPrintRenderer.write_pdf(html, base_url_resolved, out_pdf)
            except Exception as e:
                errors.append(f"WeasyPrint: {e}")
        
        # 3. Try wkhtmltopdf (fallback)
        if _HAVE_WKHTMLTOPDF:
            try:
                return WkhtmltopdfRenderer.write_pdf(html, base_url_resolved, out_pdf)
            except Exception as e:
                errors.append(f"wkhtmltopdf: {e}")
        
        # No renderers available
        error_msg = "No PDF renderers available. Install one of:\n"
        if not _HAVE_PLAYWRIGHT:
            error_msg += "1. Playwright (recommended): pip install playwright && playwright install chromium\n"
        if not _HAVE_WEASY:
            error_msg += "2. WeasyPrint: pip install weasyprint\n"
        if not _HAVE_WKHTMLTOPDF:
            error_msg += "3. wkhtmltopdf: download and install wkhtmltopdf binary\n"
        
        if errors:
            error_msg += f"\nErrors encountered: {'; '.join(errors)}"
        
        raise RuntimeError(error_msg)


# ----------------------- CLI -----------------------

def _load_json(p: str) -> Dict[str, Any]:
    return json.loads(Path(p).read_text())

def main():
    import argparse
    ap = argparse.ArgumentParser(description="Generate PDF and filled Excel for a quote.")
    ap.add_argument("--data", help="JSON payload file", default=None)
    ap.add_argument("--out", help="Output directory", default="build")
    ap.add_argument("--pdf-name", default="quote.pdf")
    ap.add_argument("--xlsx-name", default="quote.xlsx")
    ap.add_argument("--excel-template", default=None)
    ap.add_argument("--cell-map", default=None)
    ap.add_argument("--logo", default=None)
    ap.add_argument("--css", default=None)
    ap.add_argument("--template", default=None)
    ap.add_argument("--base-url", default=None)
    args = ap.parse_args()

    out_dir = Path(args.out); out_dir.mkdir(parents=True, exist_ok=True)
    if args.data:
        payload = _load_json(args.data)
    else:
        # minimal example
        payload = {
            "pdf_form_data": {
                "quotationNumber": "Q-2025-001",
                "quotationDate": "2025-08-05",
                "companyName": "ABC Manufacturing Co."
            },
            "material_information": {"material_name": "Stainless Steel", "thickness": 2.0}
        }

    cell_map = json.loads(Path(args.cell_map).read_text()) if args.cell_map else {}

    gen = DocGenerator(excel_template=args.excel_template, cell_map=cell_map)

    pdf_path = gen.render_pdf(payload, str(out_dir / args.pdf_name),
                              title="Quotation", logo_path=args.logo,
                              css_override=args.css, template_override=args.template,
                              base_url=args.base_url)
    print(f"PDF written: {pdf_path}")

    if args.excel_template and args.cell_map:
        xlsx_path = gen.fill_excel(payload, str(out_dir / args.xlsx_name))
        print(f"XLSX written: {xlsx_path}")
    else:
        print("Excel template or cell_map not supplied; skipped XLSX generation.")

if __name__ == "__main__":
    main()